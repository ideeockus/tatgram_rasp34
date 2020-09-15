from openpyxl import load_workbook

"""
A - 0 empty column
B - 1 class_name
C - 2 week_day
D -3 lesson_start_time
E -4 lesson_end_time
F - 5 subject_name
G - 6 room_number
"""

rasp_excel = load_workbook("databases/raspisanie.xlsx", read_only=True)
rasp = rasp_excel.active

# for row in rasp.rows:
#     print(row)

# print(rasp.sheetnames)

# datetime.time().hour

# for row_num in range(1, 7):
#     for cell_num

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    print(row)





