from openpyxl import load_workbook
from aiogram import types
from sqlalchemy import create_engine, MetaData, Table, Column, Integer, String, Boolean
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
import datetime
from bot_storage.configuration import postgresql_db_url, telegram_bot_token, feedback_tg_id
import requests

"""
A - 0 empty column
B - 1 class_name
C - 2 week_day
D -3 lesson_start_time
E -4 lesson_end_time
F - 5 subject_name
G - 6 room_number
"""
Base = declarative_base()  # декларативный базовый класс


class Lessons(Base):
    __tablename__ = "rasp_db"
    id = Column(Integer, primary_key=True)
    class_name = Column(String)
    week_day = Column(String)
    lesson_start_time = Column(String)
    lesson_end_time = Column(String)
    subject_name = Column(String)
    room_number = Column(String)
    teacher_name = Column(String)


# postgres_db = ""  # <ТУТ ССЫЛКА НА БД>
# if postgres_db == "":
#     postgres_db = postgresql_db_url

# engine = create_engine('sqlite:///databases/rasp.db', echo=False)
# engine = create_engine(postgresql_db_url, echo=False)
# Base.metadata.create_all(engine)
#
# db_updater_session = sessionmaker(bind=engine)
# # rasp_session = Session()


def export_xlsx_to_db(xlsx_file, updater_user_id):

    engine = create_engine(postgresql_db_url, echo=False)
    Base.metadata.create_all(engine)
    db_updater_session = sessionmaker(bind=engine)
    rasp_session = db_updater_session()

    script_start_time = datetime.datetime.now()

    rasp = None
    try:
        rasp = load_workbook(xlsx_file, read_only=True).active
    except Exception as error:
        print("Ошибка при загрузке расписания:", error)
        return

    for lsn in rasp_session.query(Lessons).all():  # очитска предыдущей таблицы
        rasp_session.delete(lsn)

    for row_num in range(3, rasp.max_row):
        class_name = rasp[row_num][1].value
        week_day = rasp[row_num][2].value
        lesson_start_time = str(rasp[row_num][3].value)
        lesson_end_time = str(rasp[row_num][4].value)
        subject_name = rasp[row_num][5].value
        room_number = rasp[row_num][6].value
        teacher_name = rasp[row_num][7].value

        if class_name is None:
            continue

        lesson = Lessons(
            class_name=class_name,
            week_day=week_day,
            lesson_start_time=lesson_start_time,
            lesson_end_time=lesson_end_time,
            subject_name=subject_name,
            room_number=room_number,
            teacher_name=teacher_name,
        )
        rasp_session.add(lesson)
        # print(class_name, week_day, lesson_start_time, subject_name, room_number, teacher_name)
        # await bot.send_chat_action(user_id, types.ChatActions.TYPING)

    # db speed test
    db_commit_start_time = datetime.datetime.now()
    xslx_processing_time = (db_commit_start_time - script_start_time).total_seconds()
    print("На обработку xslx файла ушло", xslx_processing_time, "секнуд")
    # await bot.send_message(user_id, f"На обработку xslx файла ушло {xslx_processing_time} секнуд")

    rasp_session.commit()

    # speed test
    script_end_time = datetime.datetime.now()
    db_processing_time = (script_end_time - db_commit_start_time).total_seconds()
    total_processing_time = (script_end_time - script_start_time).total_seconds()
    print("На запись в базу данных ушло", db_processing_time, "секунд")
    print("Всего затрачено", total_processing_time, "секунд")
    # await send_message_func(feedback_tg_id, f"На запись в базу данных ушло {db_processing_time} секунд")
    # await send_message_func(feedback_tg_id, f"Всего затрачено {total_processing_time} секунд")

    rasp_updated_text = "База данных расписания обновлена!\n\n" \
                           "На обработку xslx файла ушло "+str(xslx_processing_time)+" секнуд\n" \
                        "На запись в базу данных ушло "+str(db_processing_time)+" секунд\n" \
                            "Всего затрачено "+str(total_processing_time)+" секунд"

    requests.post(f"https://api.telegram.org/bot{telegram_bot_token}/sendMessage",
                  data={
                      "chat_id": updater_user_id,
                      "text": rasp_updated_text
                  })
    if updater_user_id != feedback_tg_id:
        requests.post(f"https://api.telegram.org/bot{telegram_bot_token}/sendMessage",
                      data={
                          "chat_id": feedback_tg_id,
                          "text": rasp_updated_text
                      })

# async def export_xlsx_to_db_with_msg(xlsx_file, rasp_session, message):
#     script_start_time = datetime.datetime.now()
#
#     rasp = None
#     try:
#         rasp = load_workbook(xlsx_file, read_only=True).active
#     except Exception as error:
#         print("Ошибка при загрузке расписания:", error)
#         return
#
#     for lsn in rasp_session.query(Lessons).all():  # очитска предыдущей таблицы
#         rasp_session.delete(lsn)
#
#     for row_num in range(3, rasp.max_row):
#         class_name = rasp[row_num][1].value
#         week_day = rasp[row_num][2].value
#         lesson_start_time = str(rasp[row_num][3].value)
#         lesson_end_time = str(rasp[row_num][4].value)
#         subject_name = rasp[row_num][5].value
#         room_number = rasp[row_num][6].value
#         teacher_name = rasp[row_num][7].value
#
#         if class_name is None:
#             continue
#
#         lesson = Lessons(
#             class_name=class_name,
#             week_day=week_day,
#             lesson_start_time=lesson_start_time,
#             lesson_end_time=lesson_end_time,
#             subject_name=subject_name,
#             room_number=room_number,
#             teacher_name=teacher_name,
#         )
#         rasp_session.add(lesson)
#
#     # db speed test
#     db_commit_start_time = datetime.datetime.now()
#     xslx_processing_time = (db_commit_start_time - script_start_time).total_seconds()
#     print("На обработку xslx файла ушло", xslx_processing_time, "секнуд")
#     # await bot.send_message(user_id, f"На обработку xslx файла ушло {xslx_processing_time} секнуд")
#
#     rasp_session.commit()
#
#     # speed test
#     script_end_time = datetime.datetime.now()
#     db_processing_time = (script_end_time - db_commit_start_time).total_seconds()
#     total_processing_time = (script_end_time - script_start_time).total_seconds()
#     print("На запись в базу данных ушло", db_processing_time, "секунд")
#     print("Всего затрачено", total_processing_time, "секунд")
#

# async def export_xslx_to_db_crtn(message, xlsx_file):
#     # rasp_session = Session()
#
#     # speed test
#     script_start_time = datetime.datetime.now()
#     rasp = None
#     try:
#         rasp = load_workbook(xlsx_file, read_only=True).active
#     except Exception as error:
#         print("Ошибка", error)
#         await message.answer("При загрузке расписания возникла ошибка")
#         return
#     if rasp is None:
#         await message.answer("При загрузке расписания возникла ошибка")
#         return
#
#     for lsn in rasp_session.query(Lessons).all():  # очитска предыдущей таблицы
#         rasp_session.delete(lsn)
#
#     for row_num in range(3, rasp.max_row):
#         class_name = rasp[row_num][1].value
#         week_day = rasp[row_num][2].value
#         lesson_start_time = str(rasp[row_num][3].value)
#         lesson_end_time = str(rasp[row_num][4].value)
#         subject_name = rasp[row_num][5].value
#         room_number = rasp[row_num][6].value
#         teacher_name = rasp[row_num][7].value
#
#         if class_name is None:
#             continue
#
#         lesson = Lessons(
#             class_name=class_name,
#             week_day=week_day,
#             lesson_start_time=lesson_start_time,
#             lesson_end_time=lesson_end_time,
#             subject_name=subject_name,
#             room_number=room_number,
#             teacher_name=teacher_name,
#         )
#         rasp_session.add(lesson)
#         # print(class_name, week_day, lesson_start_time, subject_name, room_number, teacher_name)
#         # await bot.send_chat_action(user_id, types.ChatActions.TYPING)
#
#     # db speed test
#     db_commit_start_time = datetime.datetime.now()
#     xslx_processing_time = (db_commit_start_time - script_start_time).total_seconds()
#     print("На обработку xslx файла ушло", xslx_processing_time, "секнуд")
#     await message.answer(f"На обработку xslx файла ушло {xslx_processing_time} секнуд")
#
#     rasp_session.commit()
#
#     # speed test
#     script_end_time = datetime.datetime.now()
#     db_processing_time = (script_end_time - db_commit_start_time).total_seconds()
#     total_processing_time = (script_end_time - script_start_time).total_seconds()
#     print("На запись в базу данных ушло", db_processing_time, "секунд")
#     print("Всего затрачено", total_processing_time, "секунд")
#     await message.answer(f"На запись в базу данных ушло {db_processing_time} секунд")
#     await message.answer(f"Всего затрачено {total_processing_time} секунд")
#

