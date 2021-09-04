from openpyxl import load_workbook
from sqlalchemy.orm import sessionmaker
from bot_storage.configuration import telegram_bot_token, feedback_tg_id
# from bot_storage.rasp_base import DbSession as dbUpdaterSession
from bot_storage import engine, Lessons
from utils import send_direct_message

import requests
import datetime



"""
A - 0 empty column
B - 1 class_name
C - 2 week_day
D - 3 lesson_start_time
E - 4 lesson_end_time
F - 5 subject_name
G - 6 room_number
"""


def export_xlsx_to_db(xlsx_file, updater_user_id):
    try:
        dbUpdaterSession = sessionmaker(bind=engine)
        rasp_session = dbUpdaterSession()

        script_start_time = datetime.datetime.now()
        rasp = load_workbook(xlsx_file, read_only=True).active

        for lsn in rasp_session.query(Lessons).all():  # очитска предыдущей таблицы
            rasp_session.delete(lsn)

        for row_num in range(3, rasp.max_row):
            class_name = rasp[row_num][1].value
            week_day = rasp[row_num][2].value
            lesson_start_time = str(rasp[row_num][3].value)
            lesson_end_time = str(rasp[row_num][4].value)
            subject_name = rasp[row_num][5].value
            room_number = rasp[row_num][6].value
            teacher_name = rasp[row_num][7].value

            if class_name is None:
                continue

            lesson = Lessons(
                class_name=class_name,
                week_day=week_day,
                lesson_start_time=lesson_start_time,
                lesson_end_time=lesson_end_time,
                subject_name=subject_name,
                room_number=room_number,
                teacher_name=teacher_name,
            )
            rasp_session.add(lesson)

        # db performance control
        db_commit_start_time = datetime.datetime.now()
        xslx_processing_time = (db_commit_start_time - script_start_time).total_seconds()
        print("На обработку xslx файла ушло", xslx_processing_time, "секнуд")
        rasp_session.commit()

        # performance control
        script_end_time = datetime.datetime.now()
        db_processing_time = (script_end_time - db_commit_start_time).total_seconds()
        total_processing_time = (script_end_time - script_start_time).total_seconds()
        print("На запись в базу данных ушло", db_processing_time, "секунд")
        print("Всего затрачено", total_processing_time, "секунд")

        rasp_updated_text = "База данных расписания обновлена!\n\n" \
                            "На обработку xslx файла ушло "+str(xslx_processing_time)+" секнуд\n" \
                            "На запись в базу данных ушло "+str(db_processing_time)+" секунд\n" \
                            "Всего затрачено "+str(total_processing_time)+" секунд"

        send_direct_message(updater_user_id, rasp_updated_text)
        if str(updater_user_id) != str(feedback_tg_id):
            send_direct_message(feedback_tg_id, rasp_updated_text)
    except Exception as e:
        print("Ошибка при загрузке расписания:", e)
        send_direct_message(updater_user_id, f"Упс! При обновлении базы произошла какая-то ошибка!\n\n{e}")
