from openpyxl import load_workbook
from sqlalchemy.orm import sessionmaker
from bot_storage import Lessons, engine

import datetime


"""
A - 0 empty column
B - 1 class_name
C - 2 week_day
D -3 lesson_start_time
E -4 lesson_end_time
F - 5 subject_name
G - 6 room_number
"""

# speed test
script_start_time = datetime.datetime.now()

DbSession = sessionmaker(bind=engine)
rasp_session = DbSession()

rasp_excel = load_workbook("databases/rasp_1_25.09.xlsx", read_only=True)
rasp = rasp_excel.active

for lsn in rasp_session.query(Lessons).all():  # очитска предыдущей таблицы
    rasp_session.delete(lsn)

for row_num in range(3, rasp.max_row):
    class_name = rasp[row_num][1].value
    week_day = rasp[row_num][2].value
    lesson_start_time = str(rasp[row_num][3].value)
    lesson_end_time = str(rasp[row_num][4].value)
    subject_name = rasp[row_num][5].value
    room_number = rasp[row_num][6].value
    teacher_name = rasp[row_num][7].value

    if class_name is None:
        continue

    lesson = Lessons(
        class_name=class_name,
        week_day=week_day,
        lesson_start_time=lesson_start_time,
        lesson_end_time=lesson_end_time,
        subject_name=subject_name,
        room_number=room_number,
        teacher_name=teacher_name,
    )
    rasp_session.add(lesson)
    print(class_name, week_day, lesson_start_time, subject_name, room_number, teacher_name)

# db speed test
db_commit_start_time = datetime.datetime.now()
print("На обработку xslx файла ушло", (db_commit_start_time-script_start_time).total_seconds(), "секнуд")

rasp_session.commit()

# speed test
script_end_time = datetime.datetime.now()
print("На запись в базу данных ушло", (script_end_time-db_commit_start_time).total_seconds(), "секунд")
print("Всего затрачено", (script_end_time-script_start_time).total_seconds(), "секунд")
